import tkinter as tk
from tkinter import ttk
import os
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Global variables
invoice_number = 1
total_rows = 0

def select_data():
    global total_rows
    filename = filedialog.askopenfilename(title="Select Data Excel File", filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
    if filename:
        # Here, you can further process the selected Excel file
        print("Selected Data File:", filename)
        selected_data_label.config(text="Selected Data File: " + filename)
        # Calculate total rows
        data_wb = load_workbook(filename)
        data_ws = data_wb.active
        total_rows = data_ws.max_row

def select_save_folder():
    foldername = filedialog.askdirectory(title="Select Folder to Save Invoices")
    if foldername:
        # Here, you can further process the selected folder
        print("Selected Folder to Save Invoices:", foldername)
        selected_save_folder_label.config(text="Selected Folder to Save Invoices: " + foldername)
        global save_folder
        save_folder = foldername

def generate_invoices():
    global invoice_number, total_rows
    template_file = "./template.xlsx"
    data_file = selected_data_label["text"].split(": ")[1]
    
    # Create folders if they don't exist
    sheets_folder = os.path.join(save_folder, "sheets")
    pdf_folder = os.path.join(save_folder, "pdf")
    os.makedirs(sheets_folder, exist_ok=True)
    os.makedirs(pdf_folder, exist_ok=True)

    # Load template
    template_wb = load_workbook(template_file)
    template_ws = template_wb.active

    # Load data
    data_wb = load_workbook(data_file)
    data_ws = data_wb.active
    
    # Load image
    logo_img = Image("./logo.png")
    logo_img.width = 270  # specify desired width
    logo_img.height = 70  # specify desired height

    # Add image to cell B1
    template_ws.add_image(logo_img, "B1")
    
    # Progress bar
    progress = ttk.Progressbar(root, orient="horizontal", length=400, mode="determinate")
    progress.pack(pady=10)

    for row_index, row in enumerate(data_ws.iter_rows(min_row=1, max_col=4, values_only=True), start=1):
        # Update progress bar
        progress["value"] = (row_index / total_rows) * 100
        progress.update_idletasks()

        # Update cell values
        template_ws['D9'] = row[1]
        template_ws['E2'] = row[0]
        template_ws['E3'] = row[0]
        template_ws['B16'] = row[2]
        template_ws['E16'] = row[3]
        template_ws['E1'] = invoice_number

        # Increment invoice number
        invoice_number += 1

        # Save the modified template to the selected folder
        invoice_filename = f"{save_folder}/sheets/Invoice_{invoice_number}.xlsx"
        template_wb.save(invoice_filename)

    print("Invoices generated successfully!")

# Create main application window
root = tk.Tk()
root.title("Invoicing System")
root.geometry("700x300")  # Set the initial size of the window

# Create label to display selected data file
selected_data_label = tk.Label(root, text="")
selected_data_label.pack(pady=5)

# Create button for selecting data
data_button = tk.Button(root, text="Select Data Excel File", command=select_data, width=20)
data_button.pack(pady=5)

# Create label to display selected folder to save invoices
selected_save_folder_label = tk.Label(root, text="")
selected_save_folder_label.pack(pady=5)

# Create button for selecting folder to save invoices
save_button = tk.Button(root, text="Save to", command=select_save_folder, width=20)
save_button.pack(pady=5)

# Create button for generating invoices
generate_button = tk.Button(root, text="Generate Invoices", command=generate_invoices, width=20)
generate_button.pack(pady=5)

# Run the application
root.mainloop()
